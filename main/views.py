from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
# from django.core.files.storage import default_storage
# from django.core.files.base import ContentFile
# import docx
# from docx.enum.text import WD_COLOR_INDEX
from openpyxl import load_workbook
# # import img2pdf
# import convertapi
import djangoProject1.firebasestr as pyr
import pdfrw
import os
from PIL import Image
# from PyPDF2 import PdfFileWriter, PdfFileReader
import fitz

os.chdir("media/")


#
# @api_view(["GET", "POST"])
# def createSgs(request):
#     import pythoncom
#     pythoncom.CoInitialize()
#     data = request.data
#     data = data.dict()
#
#     print(data)
#     # return Response(True)
#     wb = load_workbook("cemal.xlsx")
#     ws = wb.active
#     ws["B5"] = data["carBrand"]
#     ws["D5"] = data["carModel"]
#     ws["E5"] = int(data["carYear"])
#     ws["F5"] = int(data["old"])
#     ws["G5"] = data["country"]
#     ws["H5"] = data["vinNumber"]
#
#     wb.save("{}.xlsx".format(data['vinNumber']))
#     pyr.upload("{}/{}.xlsx".format(data['vinNumber'], data['vinNumber']), "{}.xlsx".format(data['vinNumber']))
#
#     document = docx.Document("2nd.docx")
#     table = document.tables[0]
#
#     # company name
#     table.column_cells(13)[7].text = data["exporterCompany"]
#
#     # adress
#     table.column_cells(13)[8].text = data["exporterAddress"]
#
#     # contact person
#     table.column_cells(13)[9].text = data["contactPerson"]
#
#     # email
#     table.column_cells(13)[10].text = data["email"]
#
#     # telephone no
#     table.column_cells(13)[11].text = data["phone"]
#
#     # Importer ***
#     # Company Name
#     table.column_cells(14)[7].text = data["importerCompany"]
#
#     # Address
#     table.column_cells(14)[8].text = data["importerAddress"]
#
#     # Invoice no / date
#     table.column_cells(20)[13].text = data["invoiceNoDate"]
#     table.column_cells(20)[13].paragraphs[0].runs[0].font.highlight_color = WD_COLOR_INDEX.YELLOW
#
#     # shading_elm_1 = parse_xml(r'<w:shd {} w:fill="faff00"/>'.format(nsdecls('w')))
#     # table.column_cells(20)[13]._tc.get_or_add_tcPr().append(shading_elm_1)
#
#     document.tables[0] = table
#     document.save("doc.docx")
#
#     convertapi.api_secret = 'gKDNm1UdZ94tL5zI'
#     files = convertapi.convert('png', {
#         'File': 'doc.docx',
#     }, from_format='docx').save_files('.')
#     my_file = 'doc.png'
#     base = os.path.splitext(my_file)[0]
#     os.rename(my_file, base + '.jpeg')
#
#     with open("{}.pdf".format(data['vinNumber']), "wb") as f:
#         f.write(
#             img2pdf.convert([i for i in os.listdir(os.curdir) if i.endswith(".jpeg")]))
#
#     sgs = Sgs(pdfFile="{}.pdf".format(data['vinNumber']), xlsxFile="{}.xlsx".format(data['vinNumber']))
#     serializer = SgsSerializer(sgs)
#     if serializer.is_valid():
#         serializer.save()
#     print(1232)
#     print(serializer.errors)
#
#     os.remove("doc.docx")
#     os.remove("doc.jpeg")
#     os.remove("doc-2.png")
#
#     pyr.upload("{}/{}.pdf".format(data['vinNumber'], data['vinNumber']), "{}.pdf".format(data['vinNumber']))
#
#     os.remove("{}.pdf".format(data['vinNumber']))
#     os.remove("{}.xlsx".format(data['vinNumber']))
#
#     Sgs()
#     SgsSerializer()
#
#     return Response(True)


@api_view(["POST"])
def createXlsx(request):
    # import pythoncom
    # pythoncom.CoInitialize()
    data = request.data
    data = data.dict()

    print(data)
    # return Response(True)
    wb = load_workbook("cemal.xlsx")
    ws = wb.active
    ws["B5"] = data["carBrand"]
    ws["D5"] = data["carModel"]
    ws["E5"] = int(data["carYear"])
    ws["F5"] = int(data["old"])
    ws["G5"] = data["country"]
    ws["H5"] = data["vinNumber"]

    wb.save("{}.xlsx".format(data['vinNumber']))
    pyr.upload("{}/{}.xlsx".format(data['vinNumber'], data['vinNumber']), "{}.xlsx".format(data['vinNumber']))
    os.remove("{}.xlsx".format(data['vinNumber']))

    return Response(True)


# @api_view(["POST"])
# def createSgs(request):
#     data = request.data.dict()
#     file = data['invoice']
#
#     filename = file.name.split(".")
#     filename[-1] = "jpeg"
#     filename = ".".join(filename)
#
#     return Response(True)
#     path = default_storage.save(filename, ContentFile(file.read()))
#
#     # shutil.move(path, "files/{}".format(filename))
#
#     import pythoncom
#     pythoncom.CoInitialize()
#     data = request.data
#     data = data.dict()
#     document = docx.Document("2nd.docx")
#     table = document.tables[0]
#
#     # company name
#     table.column_cells(13)[7].text = data["exporterCompany"]
#
#     # adress
#     table.column_cells(13)[8].text = data["exporterAddress"]
#
#     # contact person
#     table.column_cells(13)[9].text = data["contactPerson"]
#
#     # email
#     table.column_cells(13)[10].text = data["email"]
#
#     # telephone no
#     table.column_cells(13)[11].text = data["phone"]
#
#     # Importer ***
#     # Company Name
#     table.column_cells(14)[7].text = data["importerCompany"]
#
#     # Address
#     table.column_cells(14)[8].text = data["importerAddress"]
#
#     # Invoice no / date
#     table.column_cells(20)[13].text = data["invoiceNoDate"]
#     table.column_cells(20)[13].paragraphs[0].runs[0].font.highlight_color = WD_COLOR_INDEX.YELLOW
#
#     # shading_elm_1 = parse_xml(r'<w:shd {} w:fill="faff00"/>'.format(nsdecls('w')))
#     # table.column_cells(20)[13]._tc.get_or_add_tcPr().append(shading_elm_1)
#
#     document.tables[0] = table
#     document.save("doc.docx")
#
#     convertapi.api_secret = 'gKDNm1UdZ94tL5zI'  # eski
#     # convertapi.api_secret = '9R7pyBfJESrq7tJ1'  # yeni
#     files = convertapi.convert('png', {
#         'File': 'doc.docx',
#     }, from_format='docx').save_files('.')
#     my_file = 'doc.png'
#     base = os.path.splitext(my_file)[0]
#     os.rename(my_file, base + '.jpeg')
#
#     with open("{}.pdf".format(data['vinNumber']), "wb") as f:
#         f.write(
#             img2pdf.convert([i for i in os.listdir(os.curdir) if i.endswith(".jpeg")]))
#
#     os.remove("doc.docx")
#     os.remove("doc.jpeg")
#     os.remove("doc-2.png")
#
#     pyr.upload("{}/{}.pdf".format(data['vinNumber'], data['vinNumber']), "{}.pdf".format(data['vinNumber']))
#
#     os.remove("{}.pdf".format(data['vinNumber']))
#     os.remove(filename)
#
#     return Response(True)


def fill_pdf(input_pdf_path, output_pdf_path, data_dict):
    pdf_template = "sgs.pdf"
    pdf_output = "output.pdf"

    template_pdf = pdfrw.PdfReader(pdf_template)
    ANNOT_KEY = '/Annots'
    ANNOT_FIELD_KEY = '/T'
    ANNOT_VAL_KEY = '/V'
    ANNOT_RECT_KEY = '/Rect'
    SUBTYPE_KEY = '/Subtype'
    WIDGET_SUBTYPE_KEY = '/Widget'

    template_pdf = pdfrw.PdfReader(input_pdf_path)
    for page in template_pdf.pages:
        annotations = page[ANNOT_KEY]
        for annotation in annotations:
            template_pdf.Root.AcroForm.update(pdfrw.PdfDict(NeedAppearances=pdfrw.PdfObject('true')))

            if annotation[SUBTYPE_KEY] == WIDGET_SUBTYPE_KEY:
                if annotation[ANNOT_FIELD_KEY]:
                    key = annotation[ANNOT_FIELD_KEY][1:-1]
                    if key in data_dict.keys():
                        if type(data_dict[key]) == bool:
                            if data_dict[key] == True:
                                annotation.update(pdfrw.PdfDict(
                                    AS=pdfrw.PdfName('Yes')))
                        else:
                            annotation.update(
                                pdfrw.PdfDict(V='{}'.format(data_dict[key]))
                            )
                            annotation.update(pdfrw.PdfDict(AP=''))

    template_pdf.Root.AcroForm.update(pdfrw.PdfDict(NeedAppearances=pdfrw.PdfObject('true')))
    pdfrw.PdfWriter().write(output_pdf_path, template_pdf)


@api_view(["POST"])
def createSgs2(request):
    data = request.data.dict()
    print(data["type"] == "Kendal Deniz")
    # return Response(True)
    file = data['invoice']
    # if os.path.isfile("output.pdf"):
    #     os.remove("output.pdf")

    data_dict = {
        "Exporter_Company Name": data["exporterCompany"],
        "Exporter_Company Address": data["exporterAddress"],
        "Exporter_Contact Person": data["contactPerson"],
        "Exporter_EMail Address": data["email"],
        "Exporter_Telephone No": data["phone"],
        "Importer_Company Name": data["importerCompany"],
        "Importer_Company Address": data["importerAddress"],
        "Single Shipment": True,
        " & Date": data["invoiceNoDate"],
    }

    fill_pdf("sgs.pdf", "output.pdf", data_dict)

    image1 = Image.open(file)
    im1 = image1.convert('RGB')
    im1.save("image_pdf.pdf")

    if data["type"] == "Kendal Deniz":
        image2 = Image.open("ikinci.jpg")
        im2 = image2.convert('RGB')
        im2.save("ikinci.pdf")
    else:
        image2 = Image.open("yeni.jpg")
        im2 = image2.convert('RGB')
        im2.save("ikinci.pdf")

    original_pdf_path = "output.pdf"
    extra_page_path = "ikinci.pdf"
    extra_page_path2 = "image_pdf.pdf"
    output_file_path = "example-extended.pdf"

    original_pdf = fitz.open(original_pdf_path)
    extra_page = fitz.open(extra_page_path)
    extra_page2 = fitz.open(extra_page_path2)

    original_pdf.insertPDF(extra_page)
    original_pdf.insertPDF(extra_page2)
    original_pdf.save(output_file_path)

    pyr.upload("{}/{}.pdf".format(data['vinNumber'], data['vinNumber']), "example-extended.pdf")

    return Response(True)


@api_view(["GET"])
def sample(request):
    return Response({"success": True})
